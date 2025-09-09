from __future__ import annotations
import re
import datetime
import time
import pandas as pd
import zipfile
from openpyxl import load_workbook
from pathlib import Path
from typing import Tuple, Dict, Any
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_latest_ap_analysis(dir_path: str | Path) -> tuple[pd.DataFrame, Path]:
    """Load newest AP_Analysis_Report_YYYYMMDD_HHMMSS.csv from a folder.

    Args:
        dir_path (str | Path): UNC or local path to the reports folder.

    Returns:
        tuple[pd.DataFrame, Path]: The dataframe and the selected file path.
    """
    dir_path = Path(dir_path)
    ap_pattern = re.compile(r"AP_Analysis_Report_(\d{8})_(\d{6})\.csv$")
    ap_files = [f for f in dir_path.iterdir() if ap_pattern.match(f.name)]
    if not ap_files:
        raise FileNotFoundError(f"No AP_Analysis_Report_*.csv files found in {dir_path}")

    def ap_file_datetime(f: Path) -> datetime.datetime:
        m = ap_pattern.match(f.name)
        return datetime.datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")

    latest_ap_file = max(ap_files, key=ap_file_datetime)

    # Use the fast CSV engine by passing a literal delimiter (no regex)
    df = pd.read_csv(latest_ap_file, sep="^", engine="c", dtype=str)
    return df, latest_ap_file

def load_vendor_payable_workbook(
    xlsx_path: str | Path,
    *,
    read_only: bool = False,
    data_only: bool = True,
):
    """Load the Vendor Payable Report workbook via openpyxl.

    Args:
        xlsx_path (str | Path): Absolute path to the .xlsx file.
        read_only (bool): Open in read-only mode (faster, lower memory; no saving).
        data_only (bool): If True, returns cell values instead of formulas where possible.

    Returns:
        openpyxl.workbook.workbook.Workbook: Loaded workbook object.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not a valid .xlsx zip or appears corrupted.
        PermissionError: If the file is locked/open (e.g., by Excel or sync client).
    """
    xlsx_path = Path(xlsx_path)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    # Quick integrity check: .xlsx should be a valid zip
    if not zipfile.is_zipfile(xlsx_path):
        raise ValueError(
            f"Not a valid .xlsx (zip) file: {xlsx_path.name}. "
            "If this is an .xlsb or corrupted file, open and re-save as .xlsx."
        )

    try:
        wb = load_workbook(filename=str(xlsx_path), read_only=read_only, data_only=data_only)
        return wb
    except zipfile.BadZipFile as e:
        raise ValueError(
            "Corrupted .xlsx (BadZipFile). In Excel, try File → Open → Open and Repair…, "
            "then Save As a new .xlsx and point the script to that file."
        ) from e
    except PermissionError as e:
        raise PermissionError(
            "Workbook appears locked (Excel open or OneDrive/SharePoint syncing). "
            "Close Excel and/or pause sync, then try again."
        ) from e

def filter_ap_analysis(
    df: pd.DataFrame,
    *,
    amount_col: str = "Amount",
    merch_col: str = "merchType",
    category_col: str = "Category",
    keep_merch_value: str = "Merch",
    keep_category_value: str = "Home Services",
) -> pd.DataFrame:
    """Apply all AP Analysis filters and return a new DataFrame."""
    out = df.copy()

    # --- Amount to numeric ---
    if amount_col not in out.columns:
        raise KeyError(f"Expected column '{amount_col}' not found")
    amt = (
        out[amount_col].astype(str).str.strip()
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
    )
    out[amount_col] = pd.to_numeric(amt, errors="coerce")
    out = out[out[amount_col].notna() & (out[amount_col] != 0)]

    # --- merchType == 'Merch' ---
    if merch_col not in out.columns:
        raise KeyError(f"Expected column '{merch_col}' not found")
    out = out[out[merch_col].fillna("").eq(keep_merch_value)]

    # --- Category in {'Home Services', 'Brands & Retail'} (case/whitespace tolerant) ---
    if category_col not in out.columns:
        raise KeyError(f"Expected column '{category_col}' not found")
    cat_norm = (
        out[category_col]
        .fillna("")
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
        .str.casefold()
    )
    allowed = {keep_category_value.casefold(), "brands & retail"}
    out = out[cat_norm.isin(allowed)]

    return out

def aggregate_vendor_data_by_date(
    df: pd.DataFrame,
    start_date: str | pd.Timestamp,
    end_date: str | pd.Timestamp,
    *,
    date_col: str = "Date",
    amount_col: str = "Amount",
    account_col: str = "Account",
    type_col: str = "Type",
    vendor_col: str = "Name",
    category_col: str = "Category",  # <-- used to populate 'Division'
) -> pd.DataFrame:
    """Aggregate AP Analysis rows into vendor-level totals for a given date range.

    Business rules (exclusive; precedence Payment > Accrued Purchases > Bill > Adjustments):
      - Accrued Purchases: account in {21109, 21142} AND type in {Bill, Bill Credit, Item Receipt}
      - Bills            : account in {21142, 21110, 21117} AND type in {Vendor Bill, Bill Credit, Vendor Credit, Journal}
      - Payments         : account in {13150, 21110, 21117} AND type in {Bill Payment, Vendor Prepayment, Vendor Prepayment Application}
      - Adjustments      : Journal rows NOT matched by the Bill rule

    Returns:
        pd.DataFrame: ['Vendor', 'Division', 'Accrued Purchases', 'Adjustments', 'Bill', 'Payment'].
    """
    # --- 1. Validate required columns ---
    required = {date_col, amount_col, account_col, type_col, vendor_col, category_col}
    missing = required - set(df.columns)
    if missing:
        raise KeyError(f"Missing required columns: {sorted(missing)}")

    # --- 2. Parse dates and filter range (inclusive) ---
    s = pd.to_datetime(start_date).normalize()
    e = pd.to_datetime(end_date).normalize()
    if pd.isna(s) or pd.isna(e):
        raise ValueError("start_date/end_date could not be parsed.")
    if s > e:
        raise ValueError("start_date cannot be after end_date.")

    tmp = df.copy()
    tmp["_date"] = pd.to_datetime(tmp[date_col], errors="coerce").dt.normalize()
    tmp = tmp[(tmp["_date"] >= s) & (tmp["_date"] <= e)]
    if tmp.empty:
        return pd.DataFrame(columns=["Vendor", "Division", "Accrued Purchases", "Adjustments", "Bill", "Payment"])

    # --- 3. Normalize amount values to numeric ---
    amt = (
        tmp[amount_col].astype(str).str.strip()
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
    )
    tmp["_amt"] = pd.to_numeric(amt, errors="coerce").fillna(0.0)

    # --- 4. Extract leading 5-digit account code ---
    acct_code = tmp[account_col].astype(str).str.extract(r"^\s*(\d{5})", expand=False)
    tmp["_acct"] = pd.to_numeric(acct_code, errors="coerce")

    # --- 5. Canonicalize transaction type strings ---
    t = tmp[type_col].astype(str).str.strip().str.casefold().str.replace(r"\s+", " ", regex=True)
    type_norm = t.replace({
        "bill": "vendor bill",
        "vendorbill": "vendor bill",
        "vendor  bill": "vendor bill",
        "journal entry": "journal",
        "itemreceipt": "item receipt",
        "billpayment": "bill payment",
        "vendorprepayment": "vendor prepayment",
        "vendorprepayment application": "vendor prepayment application",
    }, regex=False)
    tmp["_type"] = type_norm

    # --- 6. Rule sets ---
    ACCRUED_ACCTS = {21109, 21142}
    BILL_ACCTS    = {21142, 21110, 21117}
    PAY_ACCTS     = {13150, 21110, 21117}

    ACCRUED_TYPES = {"vendor bill", "bill credit", "item receipt"}
    BILL_TYPES    = {"vendor bill", "bill credit", "vendor credit", "journal"}
    PAY_TYPES     = {"bill payment", "vendor prepayment", "vendor prepayment application"}

    # --- 7. Masks ---
    accrued_mask = tmp["_acct"].isin(ACCRUED_ACCTS) & tmp["_type"].isin(ACCRUED_TYPES)
    bill_mask    = tmp["_acct"].isin(BILL_ACCTS)    & tmp["_type"].isin(BILL_TYPES)
    pay_mask     = tmp["_acct"].isin(PAY_ACCTS)     & tmp["_type"].isin(PAY_TYPES)
    journal_adj_mask = (tmp["_type"] == "journal") & ~tmp["_acct"].isin(BILL_ACCTS)

    # --- 8. Exclusive class with precedence ---
    cls = pd.Series("other", index=tmp.index)
    cls = cls.mask(pay_mask, "payment")
    cls = cls.mask((cls == "other") & accrued_mask, "accrued")
    cls = cls.mask((cls == "other") & bill_mask, "bill")
    cls = cls.mask((cls == "other") & journal_adj_mask, "adjustments")
    tmp["_class"] = cls

    # --- 9. Per-bucket numeric columns ---
    tmp["Accrued Purchases"] = tmp["_amt"].where(tmp["_class"] == "accrued", 0.0)
    tmp["Bill"]              = tmp["_amt"].where(tmp["_class"] == "bill", 0.0)
    tmp["Payment"]           = tmp["_amt"].where(tmp["_class"] == "payment", 0.0)
    tmp["Adjustments"]       = tmp["_amt"].where(tmp["_class"] == "adjustments", 0.0)

    # --- 10. Aggregate to vendor level ---
    sums = (
        tmp.groupby(vendor_col, as_index=False)[
            ["Accrued Purchases", "Adjustments", "Bill", "Payment"]
        ].sum()
    ).rename(columns={vendor_col: "Vendor"})

    # --- 11. Bring over 'Division' from 'Category' ---
    def _pick_division(s: pd.Series):
        # most frequent non-empty value; fallback to first non-empty; else NA
        s_clean = s.dropna().astype(str).str.strip()
        if s_clean.empty:
            return pd.NA
        mode_vals = s_clean.mode()
        return mode_vals.iat[0] if not mode_vals.empty else s_clean.iloc[0]

    divisions = (
        tmp.groupby(vendor_col, as_index=False)[category_col]
           .agg(_pick_division)
           .rename(columns={vendor_col: "Vendor", category_col: "Division"})
    )

    out = sums.merge(divisions, on="Vendor", how="left")
    # Order columns
    out = out[["Vendor", "Division", "Accrued Purchases", "Adjustments", "Bill", "Payment"]]

    return out

def insert_aggregate_into_listings(
    wb: Workbook,
    df_aggregate_vendor_data: pd.DataFrame,
    *,
    sheet_name: str = "Listings",
) -> Tuple[str, int]:
    """
    Insert df_aggregate_vendor_data into the 'Listings' sheet by adding 5 columns
    (Accrued Purchases, Adjustments, Bill, Payment, Period X- Week Y) immediately
    before 'Net Change', then filling values by matching on Vendor (column B).

    Period number is computed as (current month - 1) with wrap-around (Jan -> 12).
    Week number is computed as (previous week + 1) from the latest Period column.

    The newly inserted 'Period X- Week Y' value = previous Period value
        + Accrued Purchases + Adjustments + Bill + Payment

    The newly calculated 'Net Change' = previous Period value - new Period value
    (i.e., per user instruction).

    Args:
        wb (Workbook): An *editable* openpyxl workbook (read_only must be False).
        df_aggregate_vendor_data (pd.DataFrame): Must contain columns:
            ['Vendor', 'Division', 'Accrued Purchases', 'Adjustments', 'Bill', 'Payment'].
        sheet_name (str, optional): Target worksheet name. Defaults to 'Listings'.

    Returns:
        Tuple[str, int]: (inserted_period_header, number_of_rows_updated)

    Raises:
        KeyError: If required columns are missing in either the sheet or dataframe.
        ValueError: If required headers like 'Net Change' or any 'Period X- Week Y'
                    columns cannot be found.
    """
    # --- 0. Validate inputs ---
    required_df_cols = {
        "Vendor", "Division", "Accrued Purchases", "Adjustments", "Bill", "Payment"
    }
    missing_df = required_df_cols - set(df_aggregate_vendor_data.columns)
    if missing_df:
        raise KeyError(f"df_aggregate_vendor_data missing columns: {sorted(missing_df)}")

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' not found in workbook.")
    ws: Worksheet = wb[sheet_name]

    # --- 1. Read header row (assumes header is in row 1, A1='Division') ---
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if not headers or headers[0] != "Division":
        raise KeyError("Expected 'Division' in A1 of the Listings sheet.")

    # Helper to match "Period X- Week Y"
    period_re = re.compile(r"^\s*Period\s+(\d+)\s*-\s*Week\s+(\d+)\s*$", re.I)

    # --- 2. Find the 'Net Change' column and the last Period column before it ---
    try:
        net_change_idx_1based = headers.index("Net Change") + 1  # 1-based
    except ValueError as e:
        raise ValueError("Could not find 'Net Change' header in the Listings sheet.") from e

    last_period_idx_1based = None
    last_period_label = None
    last_period_week = None

    for c in range(1, net_change_idx_1based):  # strictly before 'Net Change'
        h = headers[c - 1]
        if isinstance(h, str):
            m = period_re.match(h)
            if m:
                last_period_idx_1based = c
                last_period_label = h
                last_period_week = int(m.group(2))

    if last_period_idx_1based is None or last_period_label is None:
        raise ValueError("Could not find any 'Period X- Week Y' column before 'Net Change'.")

    # --- 3. Determine new Period and Week labels ---
    today = datetime.date.today()
    # Period = current month - 1 (wrap Jan->Dec)
    period_num = ((today.month - 2) % 12) + 1
    new_week_num = (last_period_week or 0) + 1
    new_period_label = f"Period {period_num}- Week {new_week_num}"

    # --- 4. Insert 5 columns immediately BEFORE 'Net Change' ---
    # After insertion: newly inserted columns occupy [net_change_idx_1based .. +4]
    ws.insert_cols(net_change_idx_1based, amount=5)

    col_accrued = net_change_idx_1based
    col_adjust = net_change_idx_1based + 1
    col_bill = net_change_idx_1based + 2
    col_pay = net_change_idx_1based + 3
    col_new_period = net_change_idx_1based + 4
    col_net_change = net_change_idx_1based + 5  # 'Net Change' shifts right by 5

    # Set headers for the new columns
    ws.cell(row=1, column=col_accrued,    value="Accrued Purchases")
    ws.cell(row=1, column=col_adjust,     value="Adjustments")
    ws.cell(row=1, column=col_bill,       value="Bill")
    ws.cell(row=1, column=col_pay,        value="Payment")
    ws.cell(row=1, column=col_new_period, value=new_period_label)

    # --- 5. Build a vendor lookup from the df (robust, normalized) ---
    def _norm_vendor(s: Any) -> str:
        return re.sub(r"\s+", " ", str(s).strip()).casefold()

    # Ensure numeric dtype for sums
    flows = df_aggregate_vendor_data.copy()
    for col in ["Accrued Purchases", "Adjustments", "Bill", "Payment"]:
        flows[col] = pd.to_numeric(flows[col], errors="coerce").fillna(0.0)

    vendor_map: Dict[str, Dict[str, float]] = (
        flows.set_index(flows["Vendor"].map(_norm_vendor))[
            ["Accrued Purchases", "Adjustments", "Bill", "Payment"]
        ]
        .to_dict(orient="index")
    )

    # --- 6. Helpers for numeric conversion from sheet cells ---
    def _to_float(v) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if not s:
            return 0.0
        s = s.replace(",", "")
        # Handle parentheses-negatives
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return 0.0

    # The previous period column index is unchanged by the insertion we did (we inserted at 'Net Change')
    prev_period_col = last_period_idx_1based

    # --- 7. Fill values row-by-row using Vendor match in column B ---
    updated_rows = 0
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        vendor_cell = ws.cell(row=r, column=2).value  # Column B
        if vendor_cell is None or str(vendor_cell).strip() == "":
            continue

        key = _norm_vendor(vendor_cell)
        data = vendor_map.get(key, None)

        # Pull flows; default to 0.0 if vendor not present in df
        acc_val = data["Accrued Purchases"] if data else 0.0
        adj_val = data["Adjustments"] if data else 0.0
        bill_val = data["Bill"] if data else 0.0
        pay_val = data["Payment"] if data else 0.0

        # Previous period cumulative value
        prev_period_val = _to_float(ws.cell(row=r, column=prev_period_col).value)

        # New period cumulative = previous + flows
        new_period_val = prev_period_val + acc_val + adj_val + bill_val + pay_val

        # Net Change = new - previous
        net_change_val = new_period_val - prev_period_val

        # Write values
        ws.cell(row=r, column=col_accrued,    value=acc_val).number_format = "#,##0.00"
        ws.cell(row=r, column=col_adjust,     value=adj_val).number_format = "#,##0.00"
        ws.cell(row=r, column=col_bill,       value=bill_val).number_format = "#,##0.00"
        ws.cell(row=r, column=col_pay,        value=pay_val).number_format = "#,##0.00"
        ws.cell(row=r, column=col_new_period, value=new_period_val).number_format = "#,##0.00"
        ws.cell(row=r, column=col_net_change, value=net_change_val).number_format = "#,##0.00"

        updated_rows += 1

    return new_period_label, updated_rows

# ONLY FOR DEBUGGING PURPOSES
def _normalize_for_debug(df, date_col="Date", amount_col="Amount", account_col="Account", type_col="Type"):
    d = df.copy()
    d["_date"] = pd.to_datetime(d[date_col], errors="coerce").dt.normalize()
    amt = (d[amount_col].astype(str).str.strip()
           .str.replace("$", "", regex=False)
           .str.replace(",", "", regex=False)
           .str.replace("(", "-", regex=False)
           .str.replace(")", "", regex=False))
    d["_amt"] = pd.to_numeric(amt, errors="coerce").fillna(0.0)
    acct_code = d[account_col].astype(str).str.extract(r"^\s*(\d{5})", expand=False)
    d["_acct"] = pd.to_numeric(acct_code, errors="coerce")
    t = d[type_col].astype(str).str.strip().str.casefold().str.replace(r"\s+", " ", regex=True)
    d["_type"] = t.replace({
        "bill": "vendor bill",
        "vendorbill": "vendor bill",
        "vendor  bill": "vendor bill",
        "journal entry": "journal",
        "itemreceipt": "item receipt",
        "billpayment": "bill payment",
        "vendorprepayment": "vendor prepayment",
        "vendorprepayment application": "vendor prepayment application",
    }, regex=False)
    return d


if __name__ == "__main__":
    # --- 1. Start performance timer ---
    t0 = time.perf_counter()

    # --- 2. Configure input paths and dates ---
    # 2.1 Network folder containing AP Analysis exports.
    AP_DIR = r"\\SHSNGSTFSX\shs_boomi_vol\Test\AP_Report_Files"
    # 2.2 Full path to the Vendor Payable workbook.
    VENDOR_DIR = (
        r"C:\Users\tbingha\Transform HoldCo LLC\Finance AI - Documents\Project docs\AP Financial Controls\Vendor Payable WeekXX - Prepare - Ali Mohdumair\Vendor Payable Report - DO NOT MODIFY\Vendor Payable Report.xlsx"
    )
    # 2.3 Inclusive start of aggregation window (YYYY-MM-DD).
    START_DATE = "2025-08-11"
    # 2.4 Inclusive end of aggregation window (YYYY-MM-DD).
    END_DATE = "2025-08-17"

    # --- 3. Load latest AP Analysis ---
    df_ap_analysis_raw, ap_path = load_latest_ap_analysis(AP_DIR)
    print(f"Loaded AP Analysis file: {ap_path.name}  shape={df_ap_analysis_raw.shape}")

    # --- 4. Apply filters to AP Analysis data ---
    df_ap_analysis_report = filter_ap_analysis(df_ap_analysis_raw)
    print(f"After filters: {len(df_ap_analysis_report):,} rows (from {len(df_ap_analysis_raw):,})")

    # --- 5. (Optional) Export for testing ---
    # 5.1 Uncomment to write a CSV snapshot.
    df_ap_analysis_report.to_csv("df_ap_analysis_report.csv", index=False)

    # --- 6. Aggregate vendor data for date window ---
    df_aggregate_vendor_data = aggregate_vendor_data_by_date(
        df_ap_analysis_report, start_date=START_DATE, end_date=END_DATE
    )
    print(df_aggregate_vendor_data.head())
    # 6.1 Uncomment to write a CSV snapshot.
    df_aggregate_vendor_data.to_csv("df_aggregate_vendor_data.csv", index=False)

    # --- 7. Open Vendor Payable workbook (WRITE mode) ---
    xl_vendor_payable_report = load_vendor_payable_workbook(
        VENDOR_DIR, read_only=False, data_only=True
    )
    print("Sheets:", xl_vendor_payable_report.sheetnames)

    # --- 8. Insert aggregates into 'Listings' and update Net Change ---
    new_period_header, n_rows = insert_aggregate_into_listings(
        xl_vendor_payable_report, df_aggregate_vendor_data, sheet_name="Listings"
    )
    print(f"Inserted columns through '{new_period_header}'. Updated {n_rows} rows.")

    # --- 9. Save & close workbook ---
    xl_vendor_payable_report.save(VENDOR_DIR)
    xl_vendor_payable_report.close()
    print("Saved and closed workbook.")

    # --- 10. Report elapsed time ---
    t1 = time.perf_counter()
    print(f"Done in {t1 - t0:.2f}s")



    # JUST SOME DEBUGGING CODE BELOW HERE
    # dbg = _normalize_for_debug(df_ap_analysis_report)

    # BILL_ACCTS = {21142, 21110, 21117}
    # ACCRUED_ACCTS = {21109, 21142}
    # BILL_TYPES = {"vendor bill", "bill credit", "vendor credit", "journal"}
    # ACCRUED_TYPES = {"vendor bill", "bill credit", "item receipt"}
    # PAY_ACCTS = {13150, 21110, 21117}
    # PAY_TYPES  = {"bill payment", "vendor prepayment", "vendor prepayment application"}

    # bill_like = dbg[dbg["_acct"].isin(BILL_ACCTS) & dbg["_type"].isin(BILL_TYPES)]
    # print("Bill-like rows (by rule):", len(bill_like), " Sum:", bill_like["_amt"].sum())

    # print("\nBreakdown of what is counting as Bill by _type:")
    # print(bill_like.groupby("_type")["_amt"].sum().sort_values(ascending=False))

    # print("\nHow much of Bill is actually journals?")
    # print(bill_like[bill_like["_type"]=="journal"]["_amt"].sum())

    # print("\nAny 'credit memo' rows that might be missing?")
    # credit_memo = dbg[dbg["_type"]=="credit memo"]
    # print(len(credit_memo), credit_memo["_amt"].sum())

    # print("\nRows with missing _acct (can’t classify):")
    # na_acct = dbg[dbg["_acct"].isna()]
    # print(len(na_acct), na_acct["_amt"].sum())
    # print("Top 10 raw Account strings for _acct NaN:")
    # print(na_acct["Account"].astype(str).value_counts().head(10))

    # pay_mask     = dbg["_acct"].isin(PAY_ACCTS)     & dbg["_type"].isin(PAY_TYPES)
    # accrued_mask = dbg["_acct"].isin(ACCRUED_ACCTS) & dbg["_type"].isin(ACCRUED_TYPES)
    # bill_mask    = dbg["_acct"].isin(BILL_ACCTS)    & dbg["_type"].isin(BILL_TYPES)

    # # Final Bill = bill_mask AND NOT pay_mask AND NOT accrued_mask  (matches your precedence)
    # final_bill = dbg[bill_mask & ~pay_mask & ~accrued_mask]

    # print("FINAL Bill rows (after precedence):", len(final_bill), " Sum:", final_bill["_amt"].sum())
    # print("\nFinal Bill by _type:")
    # print(final_bill.groupby("_type")["_amt"].sum().sort_values())