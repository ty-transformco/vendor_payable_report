from __future__ import annotations
import re
import datetime
import time
import pandas as pd
import zipfile
import argparse
from openpyxl import load_workbook
from pathlib import Path
from typing import Tuple, Dict, Any
from typing import Optional
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo


def load_latest_ap_analysis(dir_path: str | Path) -> tuple[pd.DataFrame, Path]:
    """Load newest AP_Analysis_Report_YYYYMMDD_HHMMSS.csv from a folder.

    Args:
        dir_path (str | Path): UNC or local path to the reports folder.

    Returns:
        tuple[pd.DataFrame, Path]: The dataframe and the selected file path.
    """
    dir_path = Path(dir_path)
    ap_pattern = re.compile(r"AP_Analysis_Report_(\d{8})_(\d{6})\.csv$")
    ap_files = [f for f in dir_path.iterdir() if ap_pattern.match(f.name)]
    if not ap_files:
        raise FileNotFoundError(f"No AP_Analysis_Report_*.csv files found in {dir_path}")

    def ap_file_datetime(f: Path) -> datetime.datetime:
        m = ap_pattern.match(f.name)
        return datetime.datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S")

    latest_ap_file = max(ap_files, key=ap_file_datetime)

    # Use the fast CSV engine by passing a literal delimiter (no regex)
    df = pd.read_csv(latest_ap_file, sep="^", engine="c", dtype=str)
    return df, latest_ap_file

def load_vendor_payable_workbook(
    xlsx_path: str | Path,
    *,
    read_only: bool = False,
    data_only: bool = True,
):
    """Load the Vendor Payable Report workbook via openpyxl.

    Args:
        xlsx_path (str | Path): Absolute path to the .xlsx file.
        read_only (bool): Open in read-only mode (faster, lower memory; no saving).
        data_only (bool): If True, returns cell values instead of formulas where possible.

    Returns:
        openpyxl.workbook.workbook.Workbook: Loaded workbook object.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file is not a valid .xlsx zip or appears corrupted.
        PermissionError: If the file is locked/open (e.g., by Excel or sync client).
    """
    xlsx_path = Path(xlsx_path)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Workbook not found: {xlsx_path}")

    # Quick integrity check: .xlsx should be a valid zip
    if not zipfile.is_zipfile(xlsx_path):
        raise ValueError(
            f"Not a valid .xlsx (zip) file: {xlsx_path.name}. "
            "If this is an .xlsb or corrupted file, open and re-save as .xlsx."
        )

    try:
        wb = load_workbook(filename=str(xlsx_path), read_only=read_only, data_only=data_only)
        return wb
    except zipfile.BadZipFile as e:
        raise ValueError(
            "Corrupted .xlsx (BadZipFile). In Excel, try File → Open → Open and Repair…, "
            "then Save As a new .xlsx and point the script to that file."
        ) from e
    except PermissionError as e:
        raise PermissionError(
            "Workbook appears locked (Excel open or OneDrive/SharePoint syncing). "
            "Close Excel and/or pause sync, then try again."
        ) from e

def filter_ap_analysis(
    df: pd.DataFrame,
    *,
    amount_col: str = "Amount",
    merch_col: str = "merchType",
    category_col: str = "Category",
    keep_merch_value: str = "Merch",
    keep_category_value: str = "Home Services",  # kept for signature compatibility; not used directly
) -> pd.DataFrame:
    """Apply AP Analysis filters; consolidate categories to 'Brands & Retail' or 'Home Services' then keep only those two."""
    out = df.copy()

    # --- Amount to numeric (robust) ---
    if amount_col not in out.columns:
        raise KeyError(f"Expected column '{amount_col}' not found")
    amt = (
        out[amount_col].astype(str).str.strip()
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
    )
    out[amount_col] = pd.to_numeric(amt, errors="coerce")
    out = out[out[amount_col].notna() & (out[amount_col] != 0)]

    # --- merchType == 'Merch' ---
    if merch_col not in out.columns:
        raise KeyError(f"Expected column '{merch_col}' not found")
    out = out[out[merch_col].fillna("").eq(keep_merch_value)]

    # --- Consolidate Category ---
    if category_col not in out.columns:
        raise KeyError(f"Expected column '{category_col}' not found")

    # normalize helper
    _norm = lambda s: re.sub(r"\s+", " ", str(s).strip()).casefold()

    # categories that should map to 'Brands & Retail'
    brands_like = {
        "a",
        "brands",
        "guam",
        "parts supply chain",
        "parts/brands",
        "parts/retail/brands",
        "retail merchandise",
        "retail",
    }

    cat_norm = out[category_col].map(_norm)

    # start with 'Other', then assign two buckets
    consolidated = pd.Series("Other", index=out.index)
    consolidated = consolidated.mask(cat_norm.eq("home services"), "Home Services")
    consolidated = consolidated.mask(cat_norm.isin(brands_like), "Brands & Retail")

    # overwrite Category with consolidated label
    out[category_col] = consolidated

    # --- Keep only 'Home Services' and 'Brands & Retail' ---
    allowed = {"home services", "brands & retail"}
    out = out[out[category_col].str.casefold().isin(allowed)]

    return out

def aggregate_vendor_data_by_date(
    df: pd.DataFrame,
    start_date: str | pd.Timestamp,
    end_date: str | pd.Timestamp,
    *,
    date_col: str = "Date",
    amount_col: str = "Amount",
    account_col: str = "Account",
    type_col: str = "Type",
    vendor_col: str = "Name",
    category_col: str = "Category",  # <-- used to populate 'Division'
) -> pd.DataFrame:
    """Aggregate AP Analysis rows into vendor-level totals for a given date range.

    Business rules (exclusive; precedence Payment > Accrued Purchases > Bill > Adjustments):
      - Accrued Purchases: account in {21109, 21142} AND type in {Bill, Bill Credit, Item Receipt}
      - Bills            : account in {21142, 21110, 21117} AND type in {Vendor Bill, Bill Credit, Vendor Credit, Journal}
      - Payments         : account in {13150, 21110, 21117} AND type in {Bill Payment, Vendor Prepayment, Vendor Prepayment Application}
      - Adjustments      : Journal rows NOT matched by the Bill rule

    Returns:
        pd.DataFrame: ['Vendor', 'Division', 'Accrued Purchases', 'Adjustments', 'Bill', 'Payment'].
    """
    # --- 1. Validate required columns ---
    required = {date_col, amount_col, account_col, type_col, vendor_col, category_col}
    missing = required - set(df.columns)
    if missing:
        raise KeyError(f"Missing required columns: {sorted(missing)}")

    # --- 2. Parse dates and filter range (inclusive) ---
    s = pd.to_datetime(start_date).normalize()
    e = pd.to_datetime(end_date).normalize()
    if pd.isna(s) or pd.isna(e):
        raise ValueError("start_date/end_date could not be parsed.")
    if s > e:
        raise ValueError("start_date cannot be after end_date.")

    tmp = df.copy()
    tmp["_date"] = pd.to_datetime(tmp[date_col], errors="coerce").dt.normalize()
    tmp = tmp[(tmp["_date"] >= s) & (tmp["_date"] <= e)]
    if tmp.empty:
        return pd.DataFrame(columns=["Vendor", "Division", "Accrued Purchases", "Adjustments", "Bill", "Payment"])

    # --- 3. Normalize amount values to numeric ---
    amt = (
        tmp[amount_col].astype(str).str.strip()
        .str.replace("$", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("(", "-", regex=False)
        .str.replace(")", "", regex=False)
    )
    tmp["_amt"] = pd.to_numeric(amt, errors="coerce").fillna(0.0)

    # --- 4. Extract leading 5-digit account code ---
    acct_code = tmp[account_col].astype(str).str.extract(r"^\s*(\d{5})", expand=False)
    tmp["_acct"] = pd.to_numeric(acct_code, errors="coerce")

    # --- 5. Canonicalize transaction type strings ---
    t = tmp[type_col].astype(str).str.strip().str.casefold().str.replace(r"\s+", " ", regex=True)
    type_norm = t.replace({
        "bill": "vendor bill",
        "vendorbill": "vendor bill",
        "vendor  bill": "vendor bill",
        "journal entry": "journal",
        "itemreceipt": "item receipt",
        "billpayment": "bill payment",
        "vendorprepayment": "vendor prepayment",
        "vendorprepayment application": "vendor prepayment application",
    }, regex=False)
    tmp["_type"] = type_norm

    # --- 6. Rule sets ---
    ACCRUED_ACCTS = {21109, 21142}
    BILL_ACCTS    = {21142, 21110, 21117}
    PAY_ACCTS     = {13150, 21110, 21117}

    ACCRUED_TYPES = {"vendor bill", "bill credit", "item receipt"}
    BILL_TYPES    = {"vendor bill", "bill credit", "vendor credit", "journal"}
    PAY_TYPES     = {"bill payment", "vendor prepayment", "vendor prepayment application"}

    # --- 7. Masks ---
    accrued_mask = tmp["_acct"].isin(ACCRUED_ACCTS) & tmp["_type"].isin(ACCRUED_TYPES)
    bill_mask    = tmp["_acct"].isin(BILL_ACCTS)    & tmp["_type"].isin(BILL_TYPES)
    pay_mask     = tmp["_acct"].isin(PAY_ACCTS)     & tmp["_type"].isin(PAY_TYPES)
    journal_adj_mask = (tmp["_type"] == "journal") & ~tmp["_acct"].isin(BILL_ACCTS)

    # --- 8. Exclusive class with precedence ---
    cls = pd.Series("other", index=tmp.index)
    cls = cls.mask(pay_mask, "payment")
    cls = cls.mask((cls == "other") & accrued_mask, "accrued")
    cls = cls.mask((cls == "other") & bill_mask, "bill")
    cls = cls.mask((cls == "other") & journal_adj_mask, "adjustments")
    tmp["_class"] = cls

    # --- 9. Per-bucket numeric columns ---
    tmp["Accrued Purchases"] = tmp["_amt"].where(tmp["_class"] == "accrued", 0.0)
    tmp["Bill"]              = tmp["_amt"].where(tmp["_class"] == "bill", 0.0)
    tmp["Payment"]           = tmp["_amt"].where(tmp["_class"] == "payment", 0.0)
    tmp["Adjustments"]       = tmp["_amt"].where(tmp["_class"] == "adjustments", 0.0)

    # --- 10. Aggregate to vendor level ---
    sums = (
        tmp.groupby(vendor_col, as_index=False)[
            ["Accrued Purchases", "Adjustments", "Bill", "Payment"]
        ].sum()
    ).rename(columns={vendor_col: "Vendor"})

    # --- 11. Bring over 'Division' from 'Category' ---
    def _pick_division(s: pd.Series):
        # most frequent non-empty value; fallback to first non-empty; else NA
        s_clean = s.dropna().astype(str).str.strip()
        if s_clean.empty:
            return pd.NA
        mode_vals = s_clean.mode()
        return mode_vals.iat[0] if not mode_vals.empty else s_clean.iloc[0]

    divisions = (
        tmp.groupby(vendor_col, as_index=False)[category_col]
           .agg(_pick_division)
           .rename(columns={vendor_col: "Vendor", category_col: "Division"})
    )

    out = sums.merge(divisions, on="Vendor", how="left")
    # Order columns
    out = out[["Vendor", "Division", "Accrued Purchases", "Adjustments", "Bill", "Payment"]]

    return out

def insert_aggregate_into_listings(
    wb: Workbook,
    df_aggregate_vendor_data: pd.DataFrame,
    *,
    sheet_name: str = "Listings",
) -> Tuple[str, int]:
    """Insert df_aggregate_vendor_data into 'Listings' by adding 5 columns before 'Net Change' and filling values."""
    import re
    import datetime
    from typing import Any, Dict
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, PatternFill

    # --- 0. Validate inputs ---
    required_df_cols = {
        "Vendor", "Division", "Accrued Purchases", "Adjustments", "Bill", "Payment"
    }
    missing_df = required_df_cols - set(df_aggregate_vendor_data.columns)
    if missing_df:
        raise KeyError(f"df_aggregate_vendor_data missing columns: {sorted(missing_df)}")

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' not found in workbook.")
    ws: Worksheet = wb[sheet_name]

    # --- 1. Read header row (assumes A1 == 'Division') ---
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if not headers or headers[0] != "Division":
        raise KeyError("Expected 'Division' in A1 of the Listings sheet.")

    period_re = re.compile(r"^\s*Period\s+(\d+)\s*-\s*Week\s+(\d+)\s*$", re.I)

    # --- 2. Find 'Net Change' and last Period column before it ---
    try:
        net_change_idx_1based = headers.index("Net Change") + 1
    except ValueError as e:
        raise ValueError("Could not find 'Net Change' header in the Listings sheet.") from e

    last_period_idx_1based = None
    last_period_label = None
    last_period_week = None
    for c in range(1, net_change_idx_1based):
        h = headers[c - 1]
        if isinstance(h, str):
            m = period_re.match(h)
            if m:
                last_period_idx_1based = c
                last_period_label = h
                last_period_week = int(m.group(2))

    if last_period_idx_1based is None or last_period_label is None:
        raise ValueError("Could not find any 'Period X- Week Y' column before 'Net Change'.")

    # --- 3. Determine new Period and Week labels ---
    today = datetime.date.today()
    period_num = ((today.month - 2) % 12) + 1  # current_month - 1, wrap Jan->Dec
    new_week_num = (last_period_week or 0) + 1
    new_period_label = f"Period {period_num}- Week {new_week_num}"

    # --- 4. Insert 5 columns immediately BEFORE 'Net Change' ---
    ws.insert_cols(net_change_idx_1based, amount=5)
    col_accrued = net_change_idx_1based
    col_adjust = net_change_idx_1based + 1
    col_bill   = net_change_idx_1based + 2
    col_pay    = net_change_idx_1based + 3
    col_new_period = net_change_idx_1based + 4
    col_net_change = net_change_idx_1based + 5  # 'Net Change' shifts right by 5

    # Set headers for the new columns
    ws.cell(row=1, column=col_accrued,    value="Accrued Purchases")
    ws.cell(row=1, column=col_adjust,     value="Adjustments")
    ws.cell(row=1, column=col_bill,       value="Bill")
    ws.cell(row=1, column=col_pay,        value="Payment")
    ws.cell(row=1, column=col_new_period, value=new_period_label)

    # --- Header fill: make ALL headers in row 1 light green ---
    header_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).fill = header_fill

    # --- 5. Build a vendor lookup from the df (robust, normalized) ---
    def _norm_vendor(s: Any) -> str:
        return re.sub(r"\s+", " ", str(s).strip()).casefold()

    flows = df_aggregate_vendor_data.copy()
    for col in ["Accrued Purchases", "Adjustments", "Bill", "Payment"]:
        flows[col] = pd.to_numeric(flows[col], errors="coerce").fillna(0.0)

    vendor_map: Dict[str, Dict[str, float]] = (
        flows.set_index(flows["Vendor"].map(_norm_vendor))[
            ["Accrued Purchases", "Adjustments", "Bill", "Payment"]
        ]
        .to_dict(orient="index")
    )

    # --- 6. Helpers ---
    def _to_float(v) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip()
        if not s:
            return 0.0
        s = s.replace(",", "")
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return 0.0

    def _write_num(row: int, col: int, val: float):
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = "#,##0.00"
        if val < 0:
            cell.font = Font(color="FFFF0000")  # red for negatives

    prev_period_col = last_period_idx_1based

    # --- 7. Fill values row-by-row using Vendor match in column B ---
    updated_rows = 0
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        vendor_cell = ws.cell(row=r, column=2).value  # Column B
        if vendor_cell is None or str(vendor_cell).strip() == "":
            continue

        key = _norm_vendor(vendor_cell)
        data = vendor_map.get(key, None)

        acc_val = data["Accrued Purchases"] if data else 0.0
        adj_val = data["Adjustments"]       if data else 0.0
        bill_val = data["Bill"]             if data else 0.0
        pay_val = data["Payment"]           if data else 0.0

        prev_period_val = _to_float(ws.cell(row=r, column=prev_period_col).value)
        new_period_val = prev_period_val + acc_val + adj_val + bill_val + pay_val

        # Net Change = new - previous (per latest instruction)
        net_change_val = new_period_val - prev_period_val

        _write_num(r, col_accrued,    acc_val)
        _write_num(r, col_adjust,     adj_val)
        _write_num(r, col_bill,       bill_val)
        _write_num(r, col_pay,        pay_val)
        _write_num(r, col_new_period, new_period_val)
        _write_num(r, col_net_change, net_change_val)

        updated_rows += 1

    return new_period_label, updated_rows

def insert_daily_aggregate_into_daily_listings(
    wb: Workbook,
    df_aggregate_vendor_data: pd.DataFrame,
    day_date: str | datetime.date | datetime.datetime,
    *,
    sheet_name: str = "Daily Listings",
) -> Tuple[str, int]:
    """
    Insert df_aggregate_vendor_data into 'Daily Listings' by adding 5 columns before 'Net Change':
      Accrued Purchases, Adjustments, Bill, Payment, MM/DD/YYYY (for the specified day).
    Uses the most-recent prior MM/DD/YYYY column as the base cumulative and adds the day's flows.

    Returns:
        (new_date_label, rows_updated)
    """
    import re
    from typing import Any, Dict
    from openpyxl.worksheet.worksheet import Worksheet

    # --- 0. Validate inputs ---
    required_df_cols = {
        "Vendor", "Division", "Accrued Purchases", "Adjustments", "Bill", "Payment"
    }
    missing_df = required_df_cols - set(df_aggregate_vendor_data.columns)
    if missing_df:
        raise KeyError(f"df_aggregate_vendor_data missing columns: {sorted(missing_df)}")

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' not found in workbook.")
    ws: Worksheet = wb[sheet_name]

    # --- 1. Headers & where 'Net Change' is ---
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not headers or headers[0] != "Division":
        raise KeyError("Expected 'Division' in A1 of the Daily Listings sheet.")

    try:
        net_change_idx_1based = headers.index("Net Change") + 1
    except ValueError as e:
        raise ValueError("Could not find 'Net Change' header in Daily Listings.") from e

    # --- 2. Find the last MM/DD/YYYY header BEFORE Net Change ---
    date_re = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")
    last_date_idx_1based = None
    last_date_label = None

    for c in range(1, net_change_idx_1based):
        h = headers[c - 1]
        # Normalize header to text for matching even if it's a real date
        if isinstance(h, (datetime.datetime, datetime.date)):
            text = h.strftime("%m/%d/%Y")
        else:
            text = str(h or "").strip()

        if date_re.match(text):
            last_date_idx_1based = c
            last_date_label = text

    if last_date_idx_1based is None:
        raise ValueError("Could not find any 'MM/DD/YYYY' date column before 'Net Change' on Daily Listings.")


    # --- 3. Determine new date header from day_date ---
    new_date = pd.to_datetime(day_date).date()
    new_date_label = new_date.strftime("%m/%d/%Y")

    # --- 4. Insert 5 columns BEFORE 'Net Change' ---
    ws.insert_cols(net_change_idx_1based, amount=5)
    col_accrued     = net_change_idx_1based
    col_adjust      = net_change_idx_1based + 1
    col_bill        = net_change_idx_1based + 2
    col_pay         = net_change_idx_1based + 3
    col_new_date    = net_change_idx_1based + 4
    col_net_change  = net_change_idx_1based + 5  # 'Net Change' shifts right by 5

    # Set headers for the new columns
    ws.cell(1, col_accrued,  "Accrued Purchases")
    ws.cell(1, col_adjust,   "Adjustments")
    ws.cell(1, col_bill,     "Bill")
    ws.cell(1, col_pay,      "Payment")
    hdr = ws.cell(1, col_new_date, new_date)  # write a real date
    hdr.number_format = "m/d/yyyy"


    # Header fill: make ALL headers light green
    header_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).fill = header_fill

    # --- 5. Build vendor lookup from the df (normalized names) ---
    def _norm_vendor(s: Any) -> str:
        return re.sub(r"\s+", " ", str(s).strip()).casefold()

    flows = df_aggregate_vendor_data.copy()
    for col in ["Accrued Purchases", "Adjustments", "Bill", "Payment"]:
        flows[col] = pd.to_numeric(flows[col], errors="coerce").fillna(0.0)

    vendor_map: Dict[str, Dict[str, float]] = (
        flows.set_index(flows["Vendor"].map(_norm_vendor))[["Accrued Purchases", "Adjustments", "Bill", "Payment"]]
             .to_dict(orient="index")
    )

    # --- 6. Helpers ---
    def _to_float(v) -> float:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "")
        if not s:
            return 0.0
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return 0.0

    def _write_num(row: int, col: int, val: float):
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = "#,##0.00"
        if val < 0:
            cell.font = Font(color="FFFF0000")  # red for negatives

    prev_date_col = last_date_idx_1based

    # --- 7. Fill values row-by-row using Vendor in column B ---
    updated_rows = 0
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        vendor_cell = ws.cell(r, 2).value
        if vendor_cell is None or str(vendor_cell).strip() == "":
            continue

        key = _norm_vendor(vendor_cell)
        data = vendor_map.get(key)

        acc_val  = data["Accrued Purchases"] if data else 0.0
        adj_val  = data["Adjustments"]       if data else 0.0
        bill_val = data["Bill"]              if data else 0.0
        pay_val  = data["Payment"]           if data else 0.0

        prev_val = _to_float(ws.cell(r, prev_date_col).value)
        new_val  = prev_val + acc_val + adj_val + bill_val + pay_val

        # Net Change = new - previous (same convention as weekly)
        net_chg = new_val - prev_val

        _write_num(r, col_accrued,  acc_val)
        _write_num(r, col_adjust,   adj_val)
        _write_num(r, col_bill,     bill_val)
        _write_num(r, col_pay,      pay_val)
        _write_num(r, col_new_date, new_val)
        _write_num(r, col_net_change, net_chg)

        updated_rows += 1

    return new_date_label, updated_rows

def extend_totals_rows(
    wb: Workbook,
    *,
    sheet_name: str = "Listings",
) -> tuple[int, int, int]:
    """
    Find the two Totals rows (Home Services and Brands & Retail) and extend their totals
    across all data columns. Totals are SUMs of the rows above:
      - Home Services totals: sum from row 2 through the row above its Totals row.
      - Brands & Retail totals: sum from the row after the Home Services totals row
        through the row above the Brands & Retail totals row.

    Also styles both Totals rows: light blue fill (#BCBEF7) and bold text.

    Returns:
        (home_services_row, brands_retail_row, num_columns_updated)
    """

    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' not found in workbook.")
    ws: Worksheet = wb[sheet_name]

    # Basic sanity: expect headers on row 1 with A1 == "Division" and B1 == "Vendor"
    if ws.cell(1, 1).value != "Division" or ws.cell(1, 2).value != "Vendor":
        raise KeyError("Expected headers in row 1 with A1='Division' and B1='Vendor'.")

    max_col = ws.max_column
    max_row = ws.max_row

    def _norm(s: Optional[str]) -> str:
        return re.sub(r"\s+", " ", str(s or "").strip()).casefold()

    # Locate Totals rows by Division="Totals" and Vendor matching the bucket name
    hs_row = None
    br_row = None
    for r in range(2, max_row + 1):
        div = _norm(ws.cell(r, 1).value)  # Column A
        ven = _norm(ws.cell(r, 2).value)  # Column B
        if div == "totals" and ven == "home services":
            hs_row = r
        elif div == "totals" and ven == "brands & retail":
            br_row = r

    if hs_row is None:
        raise ValueError("Could not find Home Services Totals row (Division='Totals', Vendor='Home Services').")
    if br_row is None:
        raise ValueError("Could not find Brands & Retail Totals row (Division='Totals', Vendor='Brands & Retail').")
    if not (2 < hs_row < br_row):
        raise ValueError(f"Unexpected Totals layout: Home Services row {hs_row}, Brands & Retail row {br_row}.")

    # Styling for Totals rows
    totals_fill = PatternFill(fill_type="solid", start_color="FFBCBEF7", end_color="FFBCBEF7")
    totals_font = Font(bold=True)

    # Helper to set SUM formula or 0 when the range would be empty
    def _set_sum_or_zero(row: int, col: int, start_row: int, end_row: int):
        cell = ws.cell(row=row, column=col)
        if start_row <= end_row:
            col_letter = get_column_letter(col)
            cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
        else:
            cell.value = 0
        cell.number_format = "#,##0.00"

    # Update totals for every *data* column (from column 3 onward).
    # Skip obviously non-numeric columns like "Notes".
    num_cols_updated = 0
    for c in range(3, max_col + 1):
        header_val = ws.cell(1, c).value
        if header_val is None or str(header_val).strip() == "":
            continue
        if isinstance(header_val, str) and _norm(header_val) == "notes":
            continue

        # Home Services Totals: sum rows 2 .. hs_row-1
        _set_sum_or_zero(hs_row, c, 2, hs_row - 1)

        # Brands & Retail Totals: sum rows hs_row+1 .. br_row-1
        _set_sum_or_zero(br_row, c, hs_row + 1, br_row - 1)

        num_cols_updated += 1

    # Style both totals rows across all columns with fill + bold
    for c in range(1, max_col + 1):
        cell_hs = ws.cell(hs_row, c)
        cell_br = ws.cell(br_row, c)
        cell_hs.fill = totals_fill
        cell_br.fill = totals_fill
        cell_hs.font = totals_font
        cell_br.font = totals_font

    return hs_row, br_row, num_cols_updated

def _find_latest_period_and_flow_cols(ws: Worksheet) -> tuple[int, list[int]]:
    """Return (latest 'Period X- Week Y' column index, [Accrued, Adjustments, Bill, Payment] cols just before it)."""
    import re
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not headers or headers[0] != "Division" or headers[1] != "Vendor":
        raise KeyError("Expected row 1 headers starting with ['Division','Vendor'].")
    try:
        net_change_idx = headers.index("Net Change") + 1
    except ValueError as e:
        raise ValueError("Could not find 'Net Change' header.") from e

    period_re = re.compile(r"^\s*Period\s+\d+\s*-\s*Week\s+\d+\s*$", re.I)
    period_col = None
    for c in range(1, net_change_idx):
        h = headers[c - 1]
        if isinstance(h, str) and period_re.match(h):
            period_col = c
    if period_col is None:
        raise ValueError("Could not locate a 'Period X- Week Y' header before 'Net Change'.")

    flow_cols = [period_col - 4, period_col - 3, period_col - 2, period_col - 1]
    expected = ["Accrued Purchases", "Adjustments", "Bill", "Payment"]
    got = [str(headers[i - 1] or "").strip() for i in flow_cols]
    if [g.casefold() for g in got] != [e.casefold() for e in expected]:
        raise ValueError(f"Unexpected flow headers before latest Period: {got}")
    return period_col, flow_cols

def _find_latest_date_and_flow_cols(ws: Worksheet) -> tuple[int, list[int]]:
    """Return (latest 'MM/DD/YYYY' col index, [Accrued, Adjustments, Bill, Payment] cols just before it)."""
    import re
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if not headers or headers[0] != "Division" or headers[1] != "Vendor":
        raise KeyError("Expected row 1 headers starting with ['Division','Vendor'].")

    try:
        net_change_idx = headers.index("Net Change") + 1
    except ValueError as e:
        raise ValueError("Could not find 'Net Change' header.") from e

    date_re = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")
    date_col = None
    # scan up to (but not including) Net Change
    for c in range(1, net_change_idx):
        h = headers[c - 1]
        if isinstance(h, (datetime.datetime, datetime.date)):
            text = h.strftime("%m/%d/%Y")
        else:
            text = str(h or "").strip()
        if date_re.match(text):
            date_col = c
    if date_col is None:
        raise ValueError("Could not locate a 'MM/DD/YYYY' header before 'Net Change'.")

    flow_cols = [date_col - 4, date_col - 3, date_col - 2, date_col - 1]
    expected = ["Accrued Purchases", "Adjustments", "Bill", "Payment"]
    got = [str(headers[i - 1] or "").strip() for i in flow_cols]
    if [g.casefold() for g in got] != [e.casefold() for e in expected]:
        raise ValueError(f"Unexpected flow headers before latest date: {got}")
    return date_col, flow_cols

def _to_float_cell(ws: Worksheet, r: int, c: int) -> float:
    """Excel-like parsing: commas, () negatives, blanks -> 0.0"""
    v = ws.cell(r, c).value
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return 0.0
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return 0.0

def extract_current_snapshot_from_listings(
    wb: Workbook, *, sheet_name: str = "Listings"
) -> pd.DataFrame:
    """
    Build a row-level snapshot from Listings with:
      ['Division','Vendor','Current Balance','Accrued Purchases','Adjustments','Bill','Payment','Net Change']
    Skips Totals rows and blank vendors.
    """
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' not found.")
    ws: Worksheet = wb[sheet_name]

    period_col, flow_cols = _find_latest_period_and_flow_cols(ws)
    col_acc, col_adj, col_bill, col_pay = flow_cols

    rows = []
    for r in range(2, ws.max_row + 1):
        division = ws.cell(r, 1).value
        vendor   = ws.cell(r, 2).value
        if not vendor or str(vendor).strip() == "":
            continue
        if isinstance(division, str) and division.strip().casefold() == "totals":
            continue

        current_balance = _to_float_cell(ws, r, period_col)
        acc = _to_float_cell(ws, r, col_acc)
        adj = _to_float_cell(ws, r, col_adj)
        bill = _to_float_cell(ws, r, col_bill)
        pay = _to_float_cell(ws, r, col_pay)
        net_change = acc + adj + bill + pay

        rows.append({
            "Division": str(division or "").strip(),
            "Vendor": str(vendor).strip(),
            "Current Balance": current_balance,
            "Accrued Purchases": acc,
            "Adjustments": adj,
            "Bill": bill,
            "Payment": pay,
            "Net Change": net_change,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["Division_norm"] = df["Division"].str.strip().str.casefold()
        df = df[df["Division_norm"].isin({"home services", "brands & retail"})].drop(columns=["Division_norm"])
    return df

def extract_current_snapshot_from_daily_listings(
    wb: Workbook, *, sheet_name: str = "Daily Listings"
) -> pd.DataFrame:
    """
    Build a snapshot from Daily Listings with:
      ['Division','Vendor','Current Balance','Accrued Purchases','Adjustments','Bill','Payment','Net Change']
    Skips Totals rows and blank vendors.
    """
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' not found.")
    ws: Worksheet = wb[sheet_name]

    date_col, flow_cols = _find_latest_date_and_flow_cols(ws)
    col_acc, col_adj, col_bill, col_pay = flow_cols

    def _to_float_cell(r: int, c: int) -> float:
        v = ws.cell(r, c).value
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "")
        if not s:
            return 0.0
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except ValueError:
            return 0.0

    rows = []
    for r in range(2, ws.max_row + 1):
        division = ws.cell(r, 1).value
        vendor   = ws.cell(r, 2).value
        if not vendor or str(vendor).strip() == "":
            continue
        if isinstance(division, str) and division.strip().casefold() == "totals":
            continue

        current_balance = _to_float_cell(r, date_col)
        acc = _to_float_cell(r, col_acc)
        adj = _to_float_cell(r, col_adj)
        bill = _to_float_cell(r, col_bill)
        pay = _to_float_cell(r, col_pay)
        net_change = acc + adj + bill + pay

        rows.append({
            "Division": str(division or "").strip(),
            "Vendor": str(vendor).strip(),
            "Current Balance": current_balance,
            "Accrued Purchases": acc,
            "Adjustments": adj,
            "Bill": bill,
            "Payment": pay,
            "Net Change": net_change,
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df["Division_norm"] = df["Division"].str.strip().str.casefold()
        df = df[df["Division_norm"].isin({"home services", "brands & retail"})].drop(columns=["Division_norm"])
    return df

def create_summary_dataframes_from_listings(
    wb: Workbook, *, sheet_name: str = "Listings", top_n: int = 5
) -> dict[str, dict[str, pd.DataFrame]]:
    """
    Build summary DataFrames from the weekly 'Listings' sheet:
      - For each Division in {'Home Services','Brands & Retail'}:
          * top_by_current_balance
          * bottom_by_current_balance
          * top_by_net_change
          * bottom_by_net_change
    """
    snap = extract_current_snapshot_from_listings(wb, sheet_name=sheet_name)

    def _agg(df_div: pd.DataFrame, metric: str) -> pd.DataFrame:
        if df_div.empty:
            return pd.DataFrame(columns=["Vendor", metric])
        return (
            df_div[["Vendor", metric]]
              .groupby("Vendor", as_index=False)[metric].sum()
        )

    def _top(df_div: pd.DataFrame, metric: str) -> pd.DataFrame:
        agg = _agg(df_div, metric)
        if agg.empty:
            return agg
        return agg.sort_values(metric, ascending=False, kind="mergesort").head(top_n).reset_index(drop=True)

    def _bottom(df_div: pd.DataFrame, metric: str) -> pd.DataFrame:
        agg = _agg(df_div, metric)
        if agg.empty:
            return agg
        # Ascending: most negative / smallest first (includes negatives)
        return agg.sort_values(metric, ascending=True, kind="mergesort").head(top_n).reset_index(drop=True)

    out: dict[str, dict[str, pd.DataFrame]] = {}
    for div in ["Home Services", "Brands & Retail"]:
        sub = snap[snap["Division"].str.strip().eq(div)]
        out[div] = {
            "top_by_current_balance":    _top(sub, "Current Balance"),
            "bottom_by_current_balance": _bottom(sub, "Current Balance"),
            "top_by_net_change":         _top(sub, "Net Change"),
            "bottom_by_net_change":      _bottom(sub, "Net Change"),
        }
    return out

def create_daily_summary_dataframes_from_daily_listings(
    wb: Workbook, *, sheet_name: str = "Daily Listings", top_n: int = 5
) -> dict[str, dict[str, pd.DataFrame]]:
    """
    Build summary DataFrames from the daily 'Daily Listings' sheet:
      - For each Division in {'Home Services','Brands & Retail'}:
          * top_by_current_balance
          * bottom_by_current_balance
          * top_by_net_change
          * bottom_by_net_change
    """
    snap = extract_current_snapshot_from_daily_listings(wb, sheet_name=sheet_name)

    def _agg(df_div: pd.DataFrame, metric: str) -> pd.DataFrame:
        if df_div.empty:
            return pd.DataFrame(columns=["Vendor", metric])
        return df_div[["Vendor", metric]].groupby("Vendor", as_index=False)[metric].sum()

    def _top(df_div: pd.DataFrame, metric: str) -> pd.DataFrame:
        agg = _agg(df_div, metric)
        if agg.empty:
            return agg
        return agg.sort_values(metric, ascending=False, kind="mergesort").head(top_n).reset_index(drop=True)

    def _bottom(df_div: pd.DataFrame, metric: str) -> pd.DataFrame:
        agg = _agg(df_div, metric)
        if agg.empty:
            return agg
        return agg.sort_values(metric, ascending=True, kind="mergesort").head(top_n).reset_index(drop=True)

    out: dict[str, dict[str, pd.DataFrame]] = {}
    for div in ["Home Services", "Brands & Retail"]:
        sub = snap[snap["Division"].str.strip().eq(div)]
        out[div] = {
            "top_by_current_balance":    _top(sub, "Current Balance"),
            "bottom_by_current_balance": _bottom(sub, "Current Balance"),
            "top_by_net_change":         _top(sub, "Net Change"),
            "bottom_by_net_change":      _bottom(sub, "Net Change"),
        }
    return out

def write_summaries_to_sheet(
    wb: Workbook,
    summaries: dict[str, dict[str, pd.DataFrame]],
    *,
    sheet_name: str = "Summary",
    start_date=None,
    end_date=None,
) -> None:
    """
    Create/overwrite `sheet_name` as the FIRST worksheet, add a Day/Week banner,
    and write paired Top/Bottom summary tables side-by-side with one blank column
    of padding between them.

    Input `summaries` structure (per Division):
        {
          "Home Services": {
            "top_by_current_balance":    DataFrame[Vendor, Current Balance],
            "bottom_by_current_balance": DataFrame[Vendor, Current Balance],
            "top_by_net_change":         DataFrame[Vendor, Net Change],
            "bottom_by_net_change":      DataFrame[Vendor, Net Change],
          },
          "Brands & Retail": { ...same keys... }
        }

    Layout:
      - Each pair is written starting at the current row.
      - LEFT table uses columns 1-2 (Vendor, Metric).
      - Column 3 is left blank as padding.
      - RIGHT table uses columns 4-5 (Vendor, Metric).
      - Numeric cells are formatted "#,##0.00".
      - Header cells use a light-green fill and bold font.
      - Columns are auto-sized per table.

    Banner:
      - If start_date and end_date are provided:
          * When start_date == end_date: the banner is just "MM/DD/YYYY".
          * Otherwise: the banner is "Week MM/DD/YYYY to MM/DD/YYYY".
      - The banner spans columns 1-5.


    The function deletes any existing sheet with `sheet_name` and recreates it at index 0.
    """

    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    import pandas as pd

    # Recreate sheet at index 0
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name, index=0)

    header_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE", end_color="FFC6EFCE")
    header_font = Font(bold=True)

    # Banner
    row_cursor = 1
    if start_date is not None and end_date is not None:
        def _fmt_us(d):
            try:
                return pd.to_datetime(d).strftime("%m/%d/%Y")
            except Exception:
                return str(d)

        same_day = pd.to_datetime(start_date).date() == pd.to_datetime(end_date).date()
        # Daily: just the date; Weekly: "Week mm/dd/yyyy to mm/dd/yyyy"
        banner = _fmt_us(start_date) if same_day else f"Week {_fmt_us(start_date)} to {_fmt_us(end_date)}"

        # Span across both tables (cols 1..5)
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=5)
        cell = ws.cell(row_cursor, 1, banner)
        cell.font = Font(bold=True, size=12)
        row_cursor += 2  # blank line after banner


    def _write_table_at(start_row: int, start_col: int, title: str, df: pd.DataFrame) -> int:
        """Write a 2-col table at (start_row, start_col). Return the last used row."""
        r = start_row
        c1, c2 = start_col, start_col + 1

        # Title
        ws.cell(r, c1, title).font = Font(bold=True)
        r += 1

        # Headers
        ws.cell(r, c1, "Vendor").fill = header_fill; ws.cell(r, c1).font = header_font
        metric_name = df.columns[1] if len(df.columns) > 1 else "Value"
        ws.cell(r, c2, metric_name).fill = header_fill; ws.cell(r, c2).font = header_font
        r += 1

        # Rows
        if not df.empty:
            for _, row in df.iterrows():
                ws.cell(r, c1, row["Vendor"])
                val = float(row.iloc[1]) if pd.notna(row.iloc[1]) else 0.0
                cell = ws.cell(r, c2, val)
                cell.number_format = "#,##0.00"
                r += 1

        # Autosize the two columns for this table
        for cc in (c1, c2):
            width = max(len(str(ws.cell(start_row + 1, cc).value or "")), 12)  # header width baseline
            for rr in range(start_row + 2, r):
                width = max(width, len(str(ws.cell(rr, cc).value or "")))
            ws.column_dimensions[get_column_letter(cc)].width = min(width + 2, 50)

        return r - 1  # last used row

    # Pair writer: left table at col=1, right table at col=4 (one col gap)
    def _write_pair(start_row: int, left_title: str, left_df: pd.DataFrame, right_title: str, right_df: pd.DataFrame) -> int:
        last_left  = _write_table_at(start_row, 1, left_title,  left_df)
        last_right = _write_table_at(start_row, 4, right_title, right_df)
        return max(last_left, last_right) + 2  # leave a blank row after the pair

    r = row_cursor
    # Home Services
    r = _write_pair(
        r,
        "Home Services — Top 5 by Current Balance",
        summaries["Home Services"]["top_by_current_balance"],
        "Home Services — Bottom 5 by Current Balance",
        summaries["Home Services"]["bottom_by_current_balance"],
    )
    r = _write_pair(
        r,
        "Home Services — Top 5 by Net Change",
        summaries["Home Services"]["top_by_net_change"],
        "Home Services — Bottom 5 by Net Change",
        summaries["Home Services"]["bottom_by_net_change"],
    )
    # Brands & Retail
    r = _write_pair(
        r,
        "Brands & Retail — Top 5 by Current Balance",
        summaries["Brands & Retail"]["top_by_current_balance"],
        "Brands & Retail — Bottom 5 by Current Balance",
        summaries["Brands & Retail"]["bottom_by_current_balance"],
    )
    _ = _write_pair(
        r,
        "Brands & Retail — Top 5 by Net Change",
        summaries["Brands & Retail"]["top_by_net_change"],
        "Brands & Retail — Bottom 5 by Net Change",
        summaries["Brands & Retail"]["bottom_by_net_change"],
    )

def _compute_auto_dates(tz: str = "America/New_York") -> tuple[str, str, str]:
    """
    Compute (START_DATE, END_DATE, DAILY_DATE) automatically.
      - DAILY_DATE = today (in the given timezone)
      - START/END  = previous week's Sunday..Saturday window
    Returns ISO strings (YYYY-MM-DD).
    """
    now_local = datetime.datetime.now(ZoneInfo(tz)).date()
    daily = now_local

    # weekday(): Mon=0..Sun=6 => days since Sunday:
    days_since_sunday = (now_local.weekday() + 1) % 7
    current_week_sunday = now_local - datetime.timedelta(days=days_since_sunday)
    prev_week_sunday = current_week_sunday - datetime.timedelta(days=7)
    prev_week_saturday = prev_week_sunday + datetime.timedelta(days=6)

    return (
        prev_week_sunday.isoformat(),
        prev_week_saturday.isoformat(),
        daily.isoformat(),
    )

def _parse_cli() -> tuple[str, str, str, bool, bool]:
    """
    Parse CLI for dates and run mode, and apply the Monday gate for weekly runs.

    DATE MODE
      AUTO (default):
        --auto-dates (or omit manual args)
        DAILY_DATE = today (in --tz)
        START/END  = previous Sunday..Saturday
        Optional: --tz America/New_York

      MANUAL:
        --manual-dates  (or provide any of --start/--end/--daily)
        Requires all three: --start --end --daily
        Flexible formats accepted; returns ISO strings.

    RUN MODE
      Default = attempt both weekly and daily.
      --only-weekly   -> run weekly steps only
      --only-daily    -> run daily steps only

    MONDAY GATE (default)
      Weekly runs only if today is Monday in --tz.
      Override with: --force-weekly
    """
    p = argparse.ArgumentParser(
        description="Post AP Analysis aggregates into Vendor Payable workbook."
    )

    # Date mode
    mode = p.add_mutually_exclusive_group(required=False)
    mode.add_argument("--auto-dates", action="store_true",
        help="Use today's date and the previous Sunday–Saturday window (default).")
    mode.add_argument("--manual-dates", action="store_true",
        help="Provide dates manually via --start, --end, --daily.")
    p.add_argument("--tz", default="America/New_York",
        help="IANA timezone for 'today' when using auto dates (default: America/New_York).")
    p.add_argument("--start", help="Manual start date (e.g., 2025-08-11 or 8/11/2025)")
    p.add_argument("--end",   help="Manual end date (e.g., 2025-08-17 or 8/17/2025)")
    p.add_argument("--daily", help="Manual daily date (e.g., 2025-08-18 or 8/18/2025)")

    # Run mode
    run_group = p.add_mutually_exclusive_group(required=False)
    run_group.add_argument("--only-weekly", action="store_true", help="Run only weekly operations.")
    run_group.add_argument("--only-daily",  action="store_true", help="Run only daily operations.")

    # Monday gate override
    p.add_argument("--force-weekly", action="store_true",
        help="Override Monday-only rule and run weekly even if today is not Monday.")

    args = p.parse_args()

    # Manual mode if explicitly chosen OR any manual date provided (unless --auto-dates also given)
    if (args.manual_dates or any([args.start, args.end, args.daily])) and not args.auto_dates:
        missing = [flag for flag, val in (("--start", args.start), ("--end", args.end), ("--daily", args.daily)) if not val]
        if missing:
            p.error(f"Manual mode requires all three: --start --end --daily (missing: {', '.join(missing)})")
        s = pd.to_datetime(args.start, errors="raise").normalize()
        e = pd.to_datetime(args.end,   errors="raise").normalize()
        d = pd.to_datetime(args.daily, errors="raise").normalize()
        if s > e:
            p.error("--start cannot be after --end")
        start_iso, end_iso, daily_iso = s.date().isoformat(), e.date().isoformat(), d.date().isoformat()
        print(f"Manual dates: start={start_iso}, end={end_iso}, daily={daily_iso}")
    else:
        start_iso, end_iso, daily_iso = _compute_auto_dates(args.tz)
        print(f"Auto dates ({args.tz}): start={start_iso}, end={end_iso}, daily={daily_iso}")

    # Initial run-mode intent
    run_weekly = not args.only_daily
    run_daily  = not args.only_weekly

    # Monday gate for weekly
    today_local = datetime.datetime.now(ZoneInfo(args.tz)).date()
    is_monday = (today_local.weekday() == 0)  # Monday=0
    if run_weekly and (not is_monday) and (not args.force_weekly):
        print(f"Weekly steps skipped: today ({today_local.isoformat()} {args.tz}) is not Monday. "
              f"Use --force-weekly to override.")
        run_weekly = False

    return start_iso, end_iso, daily_iso, run_weekly, run_daily


if __name__ == "__main__":
    # --- 1. Start performance timer ---
    t0 = time.perf_counter()

    # --- 2. Configure input paths and dates ---
    AP_DIR = r"\\SHSNGSTFSX\shs_boomi_vol\Test\AP_Report_Files"
    VENDOR_DIR = (
        r"C:\Users\tbingha\Transform HoldCo LLC\Finance AI - Documents\Project docs\AP Financial Controls\Vendor Payable WeekXX - Prepare - Ali Mohdumair\Vendor Payable Report - DO NOT MODIFY\Vendor Payable Report.xlsx"
    )

    START_DATE, END_DATE, DAILY_DATE, RUN_WEEKLY, RUN_DAILY = _parse_cli()
    print(f"Run mode -> weekly={RUN_WEEKLY}, daily={RUN_DAILY}")

    # --- 3. Load latest AP Analysis ---
    df_ap_analysis_raw, ap_path = load_latest_ap_analysis(AP_DIR)
    print(f"Loaded AP Analysis file: {ap_path.name}  shape={df_ap_analysis_raw.shape}")

    # --- 4. Apply filters ---
    df_ap_analysis_report = filter_ap_analysis(df_ap_analysis_raw)
    print(f"After filters: {len(df_ap_analysis_report):,} rows (from {len(df_ap_analysis_raw):,})")

    # --- 5. Aggregations ---
    if RUN_WEEKLY:
        df_aggregate_vendor_data = aggregate_vendor_data_by_date(
            df_ap_analysis_report, start_date=START_DATE, end_date=END_DATE
        )
        print(df_aggregate_vendor_data.head())
    else:
        df_aggregate_vendor_data = None

    if RUN_DAILY:
        df_aggregate_vendor_data_daily = aggregate_vendor_data_by_date(
            df_ap_analysis_report, start_date=DAILY_DATE, end_date=DAILY_DATE
        )
        # Optional CSV output for daily
        df_aggregate_vendor_data_daily.to_csv(
            f"df_aggregate_vendor_data_daily_{pd.to_datetime(DAILY_DATE).date().isoformat()}.csv",
            index=False, float_format="%.2f"
        )
    else:
        df_aggregate_vendor_data_daily = None

    # --- 6. Open workbook (WRITE mode) ---
    xl_vendor_payable_report = load_vendor_payable_workbook(
        VENDOR_DIR, read_only=False, data_only=True
    )
    print("Sheets:", xl_vendor_payable_report.sheetnames)

    # --- 7. Weekly operations ---
    if RUN_WEEKLY and df_aggregate_vendor_data is not None:
        new_period_header, n_rows = insert_aggregate_into_listings(
            xl_vendor_payable_report, df_aggregate_vendor_data, sheet_name="Listings"
        )
        print(f"[Weekly] Inserted columns through '{new_period_header}'. Updated {n_rows} rows.")

        hs_row, br_row, num_cols = extend_totals_rows(
            xl_vendor_payable_report, sheet_name="Listings"
        )
        print(f"[Weekly] Extended totals rows (HS row {hs_row}, BR row {br_row}) across {num_cols} columns.")

        summaries = create_summary_dataframes_from_listings(
            xl_vendor_payable_report, sheet_name="Listings", top_n=5
        )
        write_summaries_to_sheet(
            xl_vendor_payable_report,
            summaries,
            sheet_name="Weekly Summary",
            start_date=START_DATE,
            end_date=END_DATE,
        )
        print("Wrote Weekly Summary sheet with top-5 tables for both divisions.")

    # --- 8. Daily operations ---
    if RUN_DAILY and df_aggregate_vendor_data_daily is not None:
        new_date_header, n_rows_daily = insert_daily_aggregate_into_daily_listings(
            xl_vendor_payable_report, df_aggregate_vendor_data_daily, DAILY_DATE, sheet_name="Daily Listings"
        )
        print(f"[Daily] Inserted columns through '{new_date_header}'. Updated {n_rows_daily} rows.")

        hs_row_d, br_row_d, num_cols_d = extend_totals_rows(
            xl_vendor_payable_report, sheet_name="Daily Listings"
        )
        print(f"[Daily] Extended totals rows (HS row {hs_row_d}, BR row {br_row_d}) across {num_cols_d} columns.")

        daily_summaries = create_daily_summary_dataframes_from_daily_listings(
            xl_vendor_payable_report, sheet_name="Daily Listings", top_n=5
        )
        write_summaries_to_sheet(
            xl_vendor_payable_report,
            daily_summaries,
            sheet_name="Daily Summary",
            start_date=DAILY_DATE,
            end_date=DAILY_DATE,
        )
        print("Wrote Daily Summary sheet with top-5 tables for both divisions.")

    # --- 9. Save & close ---
    xl_vendor_payable_report.save(VENDOR_DIR)
    xl_vendor_payable_report.close()
    print("Saved and closed workbook.")

    # --- 10. Elapsed time ---
    t1 = time.perf_counter()
    print(f"Done in {t1 - t0:.2f}s")



